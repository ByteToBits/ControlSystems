
# export_data.py
import os
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
import openpyxl

def write_Analysis_Report(blockDataFrames: dict, blockList: list, meterList: list, 
                          outputPath: str, targetMonth: str, targetYear: str,
                          analyze_data_module):
    """
    Generate complete analysis report and save to text file.
    
    Args:
        blockDataFrames: Dictionary of block DataFrames
        blockList: List of block numbers
        meterList: List of all meter names
        outputPath: Path to output folder
        targetMonth: Target month
        targetYear: Target year
        analyze_data_module: Reference to analyze_data module for analysis functions
    """
    
    # Create output file path
    output_filename = f"Metering_Report_{targetMonth}_{targetYear}.txt"
    full_output_path = os.path.join(outputPath, "Reports", output_filename)
    
    # Ensure output directory exists
    os.makedirs(os.path.join(outputPath, "Reports"), exist_ok=True)
    
    # Open file for writing
    with open(full_output_path, 'w') as report_file:
        
        # Write header
        report_file.write("="*80 + "\n")
        report_file.write(f"METERING DATA ANALYSIS REPORT\n")
        report_file.write(f"Month: {targetMonth}/{targetYear}\n")
        report_file.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        report_file.write("="*80 + "\n\n")
        
        # Process all blocks and meters
        for block in blockList:
            
            # Find all meters in this block
            meters_in_block = [meter for meter in meterList if meter.split('_')[2] == block]
            
            # Analyze block-level statistics
            block_rt_stats = analyze_data_module.analyze_Block_RT_Data(blockDataFrames[block], meterList, block)
            block_rth_stats = analyze_data_module.analyze_Block_RTH_Data(blockDataFrames[block], meterList, block)
            
            # Write block summary
            write_Block_Statistics(report_file, block, block_rt_stats, block_rth_stats)
            
            # Write individual meter statistics
            for meter in meters_in_block:
                rt_stats = analyze_data_module.analyze_Meter_RT_Data(blockDataFrames[block], meter)
                rth_stats = analyze_data_module.analyze_Meter_RTH_Data(blockDataFrames[block], meter)
                # write_Meter_Statistics(report_file, meter, rt_stats, rth_stats)
        
        # Write footer
        report_file.write("\n" + "="*80 + "\n")
        report_file.write("END OF REPORT\n")
        report_file.write("="*80 + "\n")
    
    print(f"\nAnalysis report saved to: {full_output_path}")
    return full_output_path


def write_Meter_Statistics(file, meter_name: str, rt_stats: dict, rth_stats: dict):
    """Write meter statistics to file (internal helper)."""
    
    file.write(f"\n{'='*80}\n")
    file.write(f"METER: {meter_name}\n")
    file.write(f"{'='*80}\n")
    
    file.write(f"\n--- RT Statistics ---\n")
    file.write(f"  Totalized Value:        {rt_stats['Totalized_Value']:>15,.4f}\n")
    file.write(f"  Average Value:          {rt_stats['Average_Value']:>15,.4f}\n")
    file.write(f"  Operating Hours:        {rt_stats['Operating_Hours']:>15,.2f}\n")
    file.write(f"  Healthy Data Points:    {rt_stats['Number_of_Healthy_DataPoints']:>15,}\n")
    file.write(f"  Faulty Data Points:     {rt_stats['Number_of_Faulty_DataPoints']:>15,}\n")
    file.write(f"  Data Completeness:      {rt_stats['Data_Completeness_Percentage']:>15,.2f}%\n")
    
    file.write(f"\n--- RTH Statistics ---\n")
    file.write(f"  Monthly Consumption:    {rth_stats['Monthly_Consumption']:>15,.4f}  (BILLING)\n")
    file.write(f"  Totalized Value:        {rth_stats['Totalized_Value']:>15,.4f}\n")
    file.write(f"  First Value:            {rth_stats['First_Healthy_RTH_ProcessValue']:>15,.4f}\n")
    file.write(f"  First Timestamp:        {rth_stats['First_Healthy_RTH_Timestamp']:>15}\n")
    file.write(f"  Last Value:             {rth_stats['Last_Healthy_RTH_ProcessValue']:>15,.4f}\n")
    file.write(f"  Last Timestamp:         {rth_stats['Last_Healthy_RTH_Timestamp']:>15}\n")
    file.write(f"  Healthy Data Points:    {rth_stats['Number_of_Healthy_DataPoints']:>15,}\n")
    file.write(f"  Faulty Data Points:     {rth_stats['Number_of_Faulty_DataPoints']:>15,}\n")
    file.write(f"  Data Completeness:      {rth_stats['Data_Completeness_Percentage']:>15,.2f}%\n")


def write_Block_Statistics(file, block_number: str, block_rt_stats: dict, block_rth_stats: dict):
    """Write block statistics to file (internal helper)."""
    
    file.write(f"\n{'='*80}\n")
    file.write(f"BLOCK {block_number} - SUMMARY\n")
    file.write(f"{'='*80}\n")
    
    file.write(f"\n--- Block RT Statistics ---\n")
    file.write(f"  Number of Meters:       {block_rt_stats['Number_of_Meters']:>15}\n")
    file.write(f"  Block Totalized Value:  {block_rt_stats['Block_Totalized_Value']:>15,.4f}\n")
    file.write(f"  Block Average Value:    {block_rt_stats['Block_Average_Value']:>15,.4f}\n")
    file.write(f"  Total Operating Hours:  {block_rt_stats['Block_Total_Operating_Hours']:>15,.2f}\n")
    file.write(f"  Data Completeness:      {block_rt_stats['Block_Data_Completeness_Percentage']:>15,.2f}%\n")
    
    file.write(f"\n--- Block RTH Statistics ---\n")
    file.write(f"  Number of Meters:       {block_rth_stats['Number_of_Meters']:>15}\n")
    file.write(f"  Monthly Consumption:    {block_rth_stats['Block_Monthly_Consumption']:>15,.4f}  (BILLING)\n")
    file.write(f"  Block Totalized Value:  {block_rth_stats['Block_Totalized_Value']:>15,.4f}\n")
    file.write(f"  Data Completeness:      {block_rth_stats['Block_Data_Completeness_Percentage']:>15,.2f}%\n")


def write_DataFrames_to_Excel(blockDataFrames: dict, blockList: list, outputPath: str, targetMonth: str, targetYear: str):
    """
    Export block DataFrames to Excel file with each block as a separate sheet.
    
    Args:
        blockDataFrames: Dictionary of block DataFrames
        blockList: List of block numbers
        outputPath: Path to output folder
        targetMonth: Target month
        targetYear: Target year
    """
    
    # Create output file path
    output_filename = f"Metering_Report_{targetMonth}_{targetYear}.xlsx"
    full_output_path = os.path.join(outputPath, "Reports", output_filename)
    
    # Ensure output directory exists
    os.makedirs(os.path.join(outputPath, "Reports"), exist_ok=True)
    
    # Create Excel writer
    with pd.ExcelWriter(full_output_path, engine='openpyxl') as writer:
        
        # Write each block DataFrame to a separate sheet
        for block, df in blockDataFrames.items():
            sheet_name = f'Block {block}'
            
            # Write DataFrame starting from row 2 (leave row 1 blank)
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
        
            # General Parameters to Customize Worksheet Design
            worksheet = writer.sheets[sheet_name]
            alignment = Alignment(horizontal="center", vertical="center")
            fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
            worksheetColumnWidth = 22.5

            # Specific Parameters for Header
            header_fill = PatternFill(start_color="00153E", end_color="00153E", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            # Specific Parameters for Computed Data
            computed_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

            # Set the column widths
            for i in range(1, worksheet.max_column + 1):
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = worksheetColumnWidth
            
            # Customize Sheet Header
            for cell in worksheet[2]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment

            # Format ALL data cells (excluding header row 2)
            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
                for cell in row:
                    cell.fill = fill
                    cell.alignment = alignment

            # Apply computed column formatting
            computed_columns = ['D', 'E', 'F', 'G', 'H']
            for col_letter in computed_columns:
                for cell in worksheet[col_letter]:
                    if cell.row > 2:
                        cell.fill = computed_fill
                        cell.alignment = alignment
                
        # Create Summary Sheet
        summary_sheet = writer.book.create_sheet('Summary', 0)

        # Define alternating colors for block sections
        color1_header = PatternFill(start_color="00153E", end_color="00153E", fill_type="solid")  # Dark navy header
        color2_header = PatternFill(start_color="003B76", end_color="003B76", fill_type="solid")  # Lighter blue header
        color1_data = PatternFill(start_color="CCDAEC", end_color="CCDAEC", fill_type="solid")    # Dark tone data
        color2_data = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")    # Light tone data

        # Row 2: Merged block headers (alternating colors)
        start_col = 4
        for idx, block in enumerate(blockList):
            block_fill = color2_header if idx % 2 == 0 else color1_header 
            
            summary_sheet.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col + 2)
            cell = summary_sheet.cell(row=2, column=start_col)
            cell.value = f"Block {block}"
            cell.fill = block_fill
            cell.font = header_font
            cell.alignment = alignment
            start_col += 3

        # Row 3: Column headers (alternating colors to match blocks above)
        summary_sheet.cell(row=3, column=1, value="Timestamp").fill = color1_header
        summary_sheet.cell(row=3, column=1).font = header_font
        summary_sheet.cell(row=3, column=1).alignment = alignment

        summary_sheet.cell(row=3, column=2, value="Date").fill = color1_header
        summary_sheet.cell(row=3, column=2).font = header_font
        summary_sheet.cell(row=3, column=2).alignment = alignment

        summary_sheet.cell(row=3, column=3, value="Time").fill = color1_header
        summary_sheet.cell(row=3, column=3).font = header_font
        summary_sheet.cell(row=3, column=3).alignment = alignment

        col_index = 4
        for idx, block in enumerate(blockList):
            block_fill = color2_header if idx % 2 == 0 else color1_header 
            
            summary_sheet.cell(row=3, column=col_index, value="Total RT Sum").fill = block_fill
            summary_sheet.cell(row=3, column=col_index).font = header_font
            summary_sheet.cell(row=3, column=col_index).alignment = alignment
            
            summary_sheet.cell(row=3, column=col_index + 1, value="CWSA RT Sum").fill = block_fill
            summary_sheet.cell(row=3, column=col_index + 1).font = header_font
            summary_sheet.cell(row=3, column=col_index + 1).alignment = alignment
            
            summary_sheet.cell(row=3, column=col_index + 2, value="Retail RT Sum").fill = block_fill
            summary_sheet.cell(row=3, column=col_index + 2).font = header_font
            summary_sheet.cell(row=3, column=col_index + 2).alignment = alignment
            
            col_index += 3

        # Row 4+: Copy data with alternating colors per block section
        first_block = blockList[0]
        first_df = blockDataFrames[first_block]

        for row_idx, row in enumerate(first_df.itertuples(index=False), start=4):
            # Timestamp, Date, Time (use dark tone)
            summary_sheet.cell(row=row_idx, column=1, value=row[0]).fill = color1_data
            summary_sheet.cell(row=row_idx, column=1).alignment = alignment
            
            summary_sheet.cell(row=row_idx, column=2, value=row[1]).fill = color1_data
            summary_sheet.cell(row=row_idx, column=2).alignment = alignment
            
            summary_sheet.cell(row=row_idx, column=3, value=row[2]).fill = color1_data
            summary_sheet.cell(row=row_idx, column=3).alignment = alignment
            
            # Data columns (alternating colors per block)
            col_index = 4
            for idx, block in enumerate(blockList):
                block_df = blockDataFrames[block]
                block_row = block_df.iloc[row_idx - 4]
                
                # Alternate data colors per block
                data_fill = color2_data if idx % 2 == 0 else color1_data
                
                summary_sheet.cell(row=row_idx, column=col_index, value=block_row.iloc[3]).fill = data_fill
                summary_sheet.cell(row=row_idx, column=col_index).alignment = alignment
                
                summary_sheet.cell(row=row_idx, column=col_index + 1, value=block_row.iloc[4]).fill = data_fill
                summary_sheet.cell(row=row_idx, column=col_index + 1).alignment = alignment
                
                summary_sheet.cell(row=row_idx, column=col_index + 2, value=block_row.iloc[5]).fill = data_fill
                summary_sheet.cell(row=row_idx, column=col_index + 2).alignment = alignment
                
                col_index += 3

        # Set column widths for summary sheet
        for i in range(1, col_index):
            summary_sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18.5
        
    print(f"\nDataFrames exported to Excel: {full_output_path}")
    return full_output_path


def write_Diagnostic_Log(diagnosticsList: list, outputPath: str, targetMonth: str, targetYear: str, runtime: float):
    """
    Write diagnostic statistics to a text file in raw format.
    
    Args:
        diagnosticsList: List of diagnostic dictionaries
        outputPath: Path to output folder
        targetMonth: Target month
        targetYear: Target year
    """
    
    # Create output file path
    output_filename = f"Diagnostic_Log_{targetMonth}_{targetYear}.txt"
    full_output_path = os.path.join(outputPath, "Output Log", output_filename)
    
    # Ensure output directory exists
    os.makedirs(os.path.join(outputPath, "Output Log"), exist_ok=True)
    
    # Write to file
    with open(full_output_path, 'w') as log_file:
        
        # Write header
        log_file.write("="*80 + "\n")
        log_file.write(f"DIAGNOSTIC LOG\n")
        log_file.write(f"Month: {targetMonth}/{targetYear}\n")
        log_file.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        log_file.write(f"Total Runtime: {runtime:.2f} seconds ({runtime/60:.2f} minutes)\n")
        log_file.write("="*80 + "\n\n")
        
        # Write raw diagnostic data
        for diag in diagnosticsList:
            log_file.write(str(diag) + "\n\n")
    
    print(f"\nDiagnostic log saved to: {full_output_path}")
    return full_output_path

